import argparse
import re
import os
import sys
import glob
import shutil
from pathlib import Path
import time
from datetime import datetime, date
from dateutil import parser
import uuid
import json
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import dxpy


def get_command_line_args(arguments) -> argparse.Namespace:
    """
    Parse command line arguments

    Returns
    -------
    args : Namespace
        Namespace of command line argument inputs
    """
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--indir", "--i", help="input dir of file(s) to parse", required=True
    )
    parser.add_argument(
        "--file",
        "--f",
        nargs="+",
        help="input file(s) to parse if want to specify",
    )
    parser.add_argument(
        "--outdir",
        "--o",
        help="dir to save the output csv files",
        default=(
            "//clingen/cg/Regional Genetics Laboratories/Bioinformatics"
            "/clinvar_submission/Output/"
        ),
    )
    parser.add_argument(
        "--parsed_file_log",
        "--pf",
        help="log file to record all parsed workbook",
        default=(
            "//clingen/cg/Regional Genetics Laboratories/Bioinformatics/"
            "clinvar_submission/Output/workbooks_parsed_all_"
            "variants.txt"
        ),
    )
    parser.add_argument(
        "--clinvar_file_log",
        "--cf",
        help="log file to record all parsed workbook submitted to clinvar",
        default=(
            "//clingen/cg/Regional Genetics Laboratories/Bioinformatics/"
            "clinvar_submission/Output/workbooks_parsed_clinvar_"
            "variants.txt"
        ),
    )
    parser.add_argument(
        "--failed_file_log",
        "--ff",
        help="log file to record failed workbook",
        default=(
            "//clingen/cg/Regional Genetics Laboratories/Bioinformatics/"
            "clinvar_submission/Output/workbooks_fail_to_parse.txt"
        ),
    )
    parser.add_argument(
        "--completed_dir",
        "--cd",
        help="dir to move the successfully parsed workbooks",
        default=(
            "//clingen/cg/Regional Genetics Laboratories/Bioinformatics/"
            "clinvar_submission/Output/completed_wb/"
        ),
    )
    parser.add_argument(
        "--failed_dir",
        "--fd",
        help="dir to move the failed workbooks",
        default=(
            "//clingen/cg/Regional Genetics Laboratories/Bioinformatics/"
            "clinvar_submission/Output/failed_wb/"
        ),
    )
    parser.add_argument(
        "--subfolder",
        "--sub",
        help="subfolder in Pandora DNAnexus project",
        default=(
            "/csvs/"
        ),
    )
    parser.add_argument(
        "--unusual_sample_name",
        action="store_true",
        help="add this argument if sample name is unusual",
    )
    parser.add_argument(
        "--token", "--tk", help="DNAnexus token to log in", required=False
    )
    parser.add_argument(
        "--no_dx_upload",
        action="store_true",
        help="add this argument if don't want to upload file(s) to dx",
    )
    args = parser.parse_args(arguments)

    return args


def get_summary_fields(
    filename: str, config_variable: dict, unusual_sample_name: bool
):  # -> tuple[pd.DataFrame, str]
    """
    Extract data from summary sheet of variant workbook

    Parameters
    ----------
      variant workbook file name
      dict from config file
      boolean for unusual_sample_name

    Returns
    -------
      data frame from summary sheet
      str for error message
    """
    workbook = load_workbook(filename)
    sampleID = workbook["summary"]["B1"].value
    CI = workbook["summary"]["F1"].value
    if ";" in CI:
        split_CI = CI.split(";")
        indication = []
        Rcode = []
        for each in split_CI:
            remove_R = each.split("_")[1]
            indication.append(remove_R)
            Rcode.append(each.split("_")[0])
        new_CI = ";".join(indication)
        combined_Rcode = ";".join(Rcode)
    else:
        new_CI = CI.split("_")[1]
        combined_Rcode = CI.split("_")[0]
    panel = workbook["summary"]["F2"].value
    date_evaluated = workbook["summary"]["G22"].value
    split_sampleID = sampleID.split("-")
    instrumentID = split_sampleID[0]
    sample_ID = split_sampleID[1]
    batchID = split_sampleID[2]
    testcode = split_sampleID[3]
    probesetID = split_sampleID[5]
    ref_genome = "not_defined"
    for cell in workbook["summary"]["A"]:
        if cell.value == "Reference:":
            ref_genome = workbook["summary"][f"B{cell.row}"].value

    # checking sample naming
    error_msg = None
    if not unusual_sample_name:
        error_msg = check_sample_name(
            instrumentID, sample_ID, batchID, testcode, probesetID
        )
    d = {
        "Instrument ID": instrumentID,
        "Specimen ID": sample_ID,
        "Batch ID": batchID,
        "Test code": testcode,
        "Probeset ID": probesetID,
        "R code": combined_Rcode,
        "Preferred condition name": new_CI,
        "Panel": panel,
        "Ref genome": ref_genome,
        "Date last evaluated": date_evaluated,
    }
    df_summary = pd.DataFrame([d])

    # If no date last evaluated, use today's date
    df_summary['Date last evaluated'] = df_summary[
        'Date last evaluated'
    ].fillna(str(date.today()))

    # Catch if workbook has value for date last evaluated which is not datetime
    # compatible
    # Can test with first item in series as all rows have the same date value
    try:
        r = bool(parser.parse(str(df_summary['Date last evaluated'][0])))
    except parser._parser.ParserError:
        error_msg = (
            f"Value for date last evaluated \"{date_evaluated}\" is not "
            "compatible with datetime conversion"
        )
        return df_summary, error_msg


    df_summary["Date last evaluated"] = pd.to_datetime(
        df_summary["Date last evaluated"]
    )
    df_summary["Institution"] = config_variable["info"]["Institution"]
    df_summary["Collection method"] = config_variable["info"][
        "Collection method"
    ]
    df_summary["Allele origin"] = config_variable["info"]["Allele origin"]
    df_summary["Affected status"] = config_variable["info"]["Affected status"]

    # getting the folder name of workbook
    # the folder name should return designated folder for either CUH or NUH
    folder_name = get_folder(filename)
    if folder_name == config_variable["info"]["CUH folder"]:
        df_summary["Organisation"] = config_variable["info"][
            "CUH Organisation"
        ]

        df_summary["Organisation ID"] = config_variable["info"]["CUH org ID"]
    elif folder_name == config_variable["info"]["NUH folder"]:
        df_summary["Organisation"] = config_variable["info"][
            "NUH Organisation"
        ]

        df_summary["Organisation ID"] = config_variable["info"]["NUH org ID"]
    else:
        print("Running for the wrong folder")
        sys.exit(1)

    return df_summary, error_msg


def get_included_fields(filename: str) -> pd.DataFrame:
    """
    Extract data from included sheet of variant workbook

    Parameters
    ----------
      variant workbook file name

    Return
    ------
      data frame from included sheet
    """
    workbook = load_workbook(filename)
    num_variants = workbook["summary"]["C38"].value
    interpreted_col = get_col_letter(workbook["included"], "Interpreted")
    df = pd.read_excel(
        filename,
        sheet_name="included",
        usecols=f"A:{interpreted_col}",
        nrows=num_variants,
    )
    df_included = df[
        [
            "CHROM",
            "POS",
            "REF",
            "ALT",
            "SYMBOL",
            "HGVSc",
            "Consequence",
            "Interpreted",
            "Comment",
        ]
    ].copy()
    if len(df_included["Interpreted"].value_counts()) > 0:
        df_included["Interpreted"] = df_included["Interpreted"].str.lower()
    df_included.rename(
        columns={
            "CHROM": "Chromosome",
            "SYMBOL": "Gene symbol",
            "POS": "Start",
            "REF": "Reference allele",
            "ALT": "Alternate allele",
        },
        inplace=True,
    )
    df_included["Local ID"] = ""
    for row in range(df_included.shape[0]):
        unique_id = uuid.uuid1()
        df_included.loc[row, "Local ID"] = f"uid_{unique_id.time}"
        time.sleep(0.5)
    df_included["Linking ID"] = df_included["Local ID"]

    return df_included


def get_report_fields(
    filename: str, df_included: pd.DataFrame
):  # -> tuple[pd.DataFrame, str]
    """
    Extract data from interpret sheet(s) of variant workbook

    Parameters
    ----------
      variant workbook file name
      data frame from included sheet

    Return
    ------
      data frame from interpret sheet(s)
      str for error message

    """
    workbook = load_workbook(filename)
    field_cells = [
        ("Associated disease", "C4"),
        ("Known inheritance", "C5"),
        ("Prevalence", "C6"),
        ("HGVSc", "C3"),
        ("Germline classification", "C26"),
        ("PVS1", "H10"),
        ("PVS1_evidence", "C10"),
        ("PS1", "H11"),
        ("PS1_evidence", "C11"),
        ("PS2", "H12"),
        ("PS2_evidence", "C12"),
        ("PS3", "H13"),
        ("PS3_evidence", "C13"),
        ("PS4", "H14"),
        ("PS4_evidence", "C14"),
        ("PM1", "H15"),
        ("PM1_evidence", "C15"),
        ("PM2", "H16"),
        ("PM2_evidence", "C16"),
        ("PM3", "H17"),
        ("PM3_evidence", "C17"),
        ("PM4", "H18"),
        ("PM4_evidence", "C18"),
        ("PM5", "H19"),
        ("PM5_evidence", "C19"),
        ("PM6", "H20"),
        ("PM6_evidence", "C20"),
        ("PP1", "H21"),
        ("PP1_evidence", "C21"),
        ("PP2", "H22"),
        ("PP2_evidence", "C22"),
        ("PP3", "H23"),
        ("PP3_evidence", "C23"),
        ("PP4", "H24"),
        ("PP4_evidence", "C24"),
        ("BS1", "K16"),
        ("BS1_evidence", "C16"),
        ("BS2", "K12"),
        ("BS2_evidence", "C12"),
        ("BS3", "K13"),
        ("BS3_evidence", "C13"),
        ("BA1", "K9"),
        ("BA1_evidence", "C9"),
        ("BP2", "K17"),
        ("BP2_evidence", "C17"),
        ("BP3", "K18"),
        ("BP3_evidence", "C18"),
        ("BS4", "K21"),
        ("BS4_evidence", "C21"),
        ("BP1", "K22"),
        ("BP1_evidence", "C22"),
        ("BP4", "K23"),
        ("BP4_evidence", "C23"),
        ("BP5", "K24"),
        ("BP5_evidence", "C24"),
        ("BP7", "K25"),
        ("BP7_evidence", "C25"),
    ]
    col_name = [i[0] for i in field_cells]
    df_report = pd.DataFrame(columns=col_name)
    report_sheets = [
        idx
        for idx in workbook.sheetnames
        if idx.lower().startswith("interpret")
    ]

    for idx, sheet in enumerate(report_sheets):
        for field, cell in field_cells:
            if workbook[sheet][cell].value is not None:
                df_report.loc[idx, field] = workbook[sheet][cell].value
    df_report.reset_index(drop=True, inplace=True)
    error_msg = None
    if not df_report.empty:
        error_msg = check_interpret_table(df_report, df_included)
    if not error_msg:
        # put strength as nan if it is 'NA'
        for row in range(df_report.shape[0]):
            for column in range(5, df_report.shape[1], 2):
                if df_report.iloc[row, column] == "NA":
                    df_report.iloc[row, column] = np.nan

        # removing evidence value if no strength
        for row in range(df_report.shape[0]):
            for column in range(5, df_report.shape[1], 2):
                if df_report.isnull().iloc[row, column]:
                    df_report.iloc[row, column + 1] = np.nan

        # getting comment on classification for clinvar submission
        matched_strength = [
            ("PVS", "Very Strong"),
            ("PS", "Strong"),
            ("PM", "Moderate"),
            ("PP", "Supporting"),
            ("BS", "Supporting"),
            ("BA", "Stand-Alone"),
            ("BP", "Supporting"),
        ]
        df_report["Comment on classification"] = ""
        for row in range(df_report.shape[0]):
            evidence = []
            for column in range(5, df_report.shape[1] - 1, 2):
                if not df_report.isnull().iloc[row, column]:
                    evidence.append(
                        [
                            df_report.columns[column],
                            df_report.iloc[row, column],
                        ]
                    )
            for index, value in enumerate(evidence):
                for st1, st2 in matched_strength:
                    if st1 in value[0] and st2 == value[1]:
                        evidence[index][1] = ""
            evidence_pair = []
            for e in evidence:
                evidence_pair.append("_".join(e).rstrip("_"))
            comment_on_classification = ",".join(evidence_pair)
            df_report.iloc[
                row, df_report.columns.get_loc("Comment on classification")
            ] = comment_on_classification

    return df_report, error_msg


def check_sample_name(
    instrumentID: str,
    sample_ID: str,
    batchID: str,
    testcode: str,
    probesetID: str,
) -> str:
    """
    checking if individual parts of sample name have
    expected naming format

    Parameters
    ----------
      str values for instrumentID, sample_ID, batchID, testcode,
      probesetID

    Return
    ------
      str for error message
    """
    try:
        assert re.match(
            r"^\d{9}$", instrumentID
        ), "Unusual name for instrumentID"
        assert re.match(r"^\d{5}[A-Z]\d{4}$", sample_ID), "Unusual sampleID"
        assert re.match(r"^\d{2}[A-Z]{5}\d{1,}$", batchID), "Unusual batchID"
        assert re.match(r"^\d{4}$", testcode), "Unusual testcode"
        assert 0 < len(probesetID) < 20, "probesetID is too long/short"
        assert (
            probesetID.isalnum() and not probesetID.isalpha()
        ), "Unusual probesetID"
        error_msg = None
    except AssertionError as msg:
        error_msg = str(msg)
        print(msg)

    return error_msg


def checking_sheets(filename: str) -> str:
    """
    check if extra row(s)/col(s) are added in the sheets

    Parameters
    ----------
      variant workbook file name

    Return
    ------
      str for error message
    """
    workbook = load_workbook(filename)
    summary = workbook["summary"]
    reports = [
        idx
        for idx in workbook.sheetnames
        if idx.lower().startswith("interpret")
    ]
    try:
        assert (
            summary["G21"].value == "Date"
        ), "extra col(s) added or change(s) done in summary sheet"
        for sheet in reports:
            report = workbook[sheet]
            assert report["B26"].value == "FINAL ACMG CLASSIFICATION", (
                "extra row(s) or col(s) added or change(s) done in "
                "interpret sheet"
            )
            assert report["L8"].value == "B_POINTS", (
                "extra row(s) or col(s) added or change(s) done in "
                "interpret sheet"
            )
        error_msg = None
    except AssertionError as msg:
        error_msg = str(msg)
        print(msg)

    return error_msg


def get_col_letter(worksheet: object, col_name: str) -> str:
    """
    Getting the column letter with specific col name

    Parameters
    ----------
    openpyxl object of current sheet
    str for name of column to get col letter

    Return
    -------
        str for column letter for specific column name
    """
    col_letter = None
    for column_cell in worksheet.iter_cols(1, worksheet.max_column):
        if column_cell[0].value == col_name:
            col_letter = column_cell[0].column_letter

    return col_letter


def write_txt_file(txt_file_name: str, filename: str, msg: str) -> None:
    """
    write txt file output

    Parameters
    ----------
      str for output txt file name
      variant workbook file name
      str for error message
    """
    with open(txt_file_name, "a") as file:
        dt = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        file.write(dt + "\t " + filename + "\t " + msg + "\n")
        file.close()


def check_interpret_table(
    df_report: pd.DataFrame, df_included: pd.DataFrame
) -> str:
    """
    check if ACMG classification and HGVSc are correctly
    filled in in the interpret table(s)

    Parameters
    ----------
      df from interpret sheet(s)
      df from included sheet

    Return
    ------
      str for error message
    """
    error_msg = []
    strength_dropdown = [
        "Very Strong",
        "Strong",
        "Moderate",
        "Supporting",
        "NA",
    ]
    BA1_dropdown = [
        "Stand-Alone",
        "Very Strong",
        "Strong",
        "Moderate",
        "Supporting",
        "NA",
    ]
    for row in range(df_report.shape[0]):
        try:
            assert (
                df_report.loc[row, "Germline classification"] is not np.nan
            ), "empty ACMG classification in interpret table"
            assert df_report.loc[row, "Germline classification"] in [
                "Pathogenic",
                "Likely Pathogenic",
                "Uncertain Significance",
                "Likely Benign",
                "Benign",
            ], "wrong ACMG classification in interpret table"
            assert (
                df_report.loc[row, "HGVSc"] is not np.nan
            ), "empty HGVSc in interpret table"
            assert df_report.loc[row, "HGVSc"] in list(df_included["HGVSc"]), (
                "HGVSc in interpret table does not match with that in "
                "included sheet"
            )
            criteria_list = [
                "PVS1",
                "PS1",
                "PS2",
                "PS3",
                "PS4",
                "PM1",
                "PM2",
                "PM3",
                "PM4",
                "PM5",
                "PM6",
                "PP1",
                "PP2",
                "PP3",
                "PP4",
                "BS2",
                "BS3",
                "BS1",
                "BP2",
                "BP3",
                "BS4",
                "BP1",
                "BP4",
                "BP5",
                "BP7",
            ]
            for criteria in criteria_list:
                if df_report.loc[row, criteria] is not np.nan:
                    assert (
                        df_report.loc[row, criteria] in strength_dropdown
                    ), f"Wrong strength in {criteria}"

            if df_report.loc[row, "BA1"] is not np.nan:
                assert (
                    df_report.loc[row, "BA1"] in BA1_dropdown
                ), "Wrong strength in BA1"

        except AssertionError as msg:
            error_msg.append(str(msg))
            print(msg)
    error_msg = "".join(error_msg)

    return error_msg


def check_interpreted_col(df: pd.DataFrame) -> str:
    """
    check if interpreted col in included sheet
    is correctly filled in

    Parameters
    ----------
    merged df

    Return
    ------
      str for error message
    """
    row_yes = df[df["Interpreted"] == "yes"].index.tolist()
    error_msg = []
    for row in range(df.shape[0]):
        if row in row_yes:
            try:
                assert (
                    df.loc[row, "Germline classification"] is not np.nan
                ), f"Wrong interpreted column in row {row+1} of included sheet"
            except AssertionError as msg:
                error_msg.append(str(msg))
                print(msg)
        else:
            try:
                assert df.loc[row, "Interpreted"] == "no", (
                    f"Wrong interpreted column dropdown in row {row+1} "
                    "of included sheet"
                )
                assert (
                    df.loc[row, "Germline classification"] is np.nan
                ), f"Wrong interpreted column in row {row+1} of included sheet"
            except AssertionError as msg:
                error_msg.append(str(msg))
                print(msg)
    error_msg = " ".join(error_msg)

    return error_msg


def get_folder(filename: str) -> str:
    """
    get the folder of input file
    Parameters:
    ----------
      str for input filename

    Return:
      str for folder name
    """
    folder = os.path.basename(os.path.normpath(os.path.dirname(filename)))
    print(folder)
    return folder


def get_parsed_list(file: str) -> list:
    """
    getting the list of previously parsed workbook

    Parameters
    ----------
    str for ref file that records previously parsed workbook

    Return
    ------
    a list of previously parsed workbook
    """
    f = open(file, "r")
    lines = f.readlines()
    parsed_list = []
    for x in lines:
        columns = x.split("\t ")
        file_path = Path(columns[1])
        parsed_list.append(Path(file_path).stem + ".xlsx")
    f.close()

    return parsed_list


def check_and_create_folder(dir: str) -> None:
    """
    check if a dir exists and create
    Parameters
    ----------
    str for directory
    """
    if not os.path.exists(dir):
        os.makedirs(dir)


def dx_login(token: str) -> bool:
    """
    Function to check dxpy login
    Input: dxpy token
    Output: boolean
    """

    try:
        DX_SECURITY_CONTEXT = {
            "auth_token_type": "Bearer",
            "auth_token": str(token),
        }

        dxpy.set_security_context(DX_SECURITY_CONTEXT)
        dxpy.api.system_whoami()
        return True

    except dxpy.exceptions.InvalidAuthentication as e:
        print(e)

        return False


def main():
    arguments = get_command_line_args(sys.argv[1:])
    input_dir = arguments.indir
    if arguments.file:
        input_file = []
        for idx, file in enumerate(arguments.file):
            input_file.append(glob.glob(input_dir + file)[0])
    else:
        input_file = glob.glob(input_dir + "*.xlsx")
    if len(input_file) == 0:
        print("Input file(s) not exist")
    check_and_create_folder(arguments.outdir)
    check_and_create_folder(arguments.completed_dir)
    check_and_create_folder(arguments.failed_dir)
    if not os.path.isfile(arguments.parsed_file_log):
        with open(arguments.parsed_file_log, "w") as file:
            file.close()
    unusual_sample_name = arguments.unusual_sample_name
    no_dx_upload = arguments.no_dx_upload
    clinvar_count = 0
    with open("parser_config.json") as f:
        config_variable = json.load(f)
    parsed_list = get_parsed_list(arguments.parsed_file_log)
    # extract fields from variant workbooks as df and merged
    for filename in input_file:
        print("Running", filename)
        if (Path(filename).stem + ".xlsx") in parsed_list:
            print(filename, "is already parsed")
            continue
        error_msg_sheet = checking_sheets(filename)
        if error_msg_sheet:
            write_txt_file(
                arguments.failed_file_log, filename, error_msg_sheet
            )
            shutil.move(filename, arguments.failed_dir)
            continue
        df_summary, error_msg_name = get_summary_fields(
            filename, config_variable, unusual_sample_name
        )
        if error_msg_name:
            write_txt_file(arguments.failed_file_log, filename, error_msg_name)
            shutil.move(filename, arguments.failed_dir)
            continue
        df_included = get_included_fields(filename)
        if df_included["Interpreted"].isna().sum() != 0:
            print("Interpreted column in included sheet needs to be fixed")
            write_txt_file(
                arguments.failed_file_log,
                filename,
                "Interpreted column in included sheet needs to be fixed",
            )
            shutil.move(filename, arguments.failed_dir)
            continue
        df_report, error_msg_table = get_report_fields(filename, df_included)
        if error_msg_table:
            write_txt_file(
                arguments.failed_file_log,
                filename,
                error_msg_table,
            )
            shutil.move(filename, arguments.failed_dir)
            continue
        if not df_included.empty:
            df_merged = pd.merge(df_included, df_summary, how="cross")
            empty_workbook = False
        else:
            df_merged = pd.concat([df_summary, df_included], axis=1)
            empty_workbook = True
        df_final = pd.merge(df_merged, df_report, on="HGVSc", how="left")
        error_msg_interpreted = None
        if not empty_workbook:
            error_msg_interpreted = check_interpreted_col(df_final)
        if error_msg_interpreted:
            write_txt_file(
                arguments.failed_file_log,
                filename,
                error_msg_interpreted,
            )
            shutil.move(filename, arguments.failed_dir)
            continue
        df_final = df_final[
            [
                "Local ID",
                "Linking ID",
                "Organisation ID",
                "Gene symbol",
                "Chromosome",
                "Start",
                "Reference allele",
                "Alternate allele",
                "R code",
                "Preferred condition name",
                "Germline classification",
                "Date last evaluated",
                "Comment on classification",
                "Collection method",
                "Allele origin",
                "Affected status",
                "HGVSc",
                "Consequence",
                "Interpreted",
                "Comment",
                "Instrument ID",
                "Specimen ID",
                "Batch ID",
                "Test code",
                "Probeset ID",
                "Panel",
                "Ref genome",
                "Organisation",
                "Institution",
                "Associated disease",
                "Known inheritance",
                "Prevalence",
                "PVS1",
                "PVS1_evidence",
                "PS1",
                "PS1_evidence",
                "PS2",
                "PS2_evidence",
                "PS3",
                "PS3_evidence",
                "PS4",
                "PS4_evidence",
                "PM1",
                "PM1_evidence",
                "PM2",
                "PM2_evidence",
                "PM3",
                "PM3_evidence",
                "PM4",
                "PM4_evidence",
                "PM5",
                "PM5_evidence",
                "PM6",
                "PM6_evidence",
                "PP1",
                "PP1_evidence",
                "PP2",
                "PP2_evidence",
                "PP3",
                "PP3_evidence",
                "PP4",
                "PP4_evidence",
                "BS1",
                "BS1_evidence",
                "BS2",
                "BS2_evidence",
                "BS3",
                "BS3_evidence",
                "BA1",
                "BA1_evidence",
                "BP2",
                "BP2_evidence",
                "BP3",
                "BP3_evidence",
                "BS4",
                "BS4_evidence",
                "BP1",
                "BP1_evidence",
                "BP4",
                "BP4_evidence",
                "BP5",
                "BP5_evidence",
                "BP7",
                "BP7_evidence",
            ]
        ]
        if empty_workbook:
            df_final.fillna("null", inplace=True)
        else:
            df_final["Germline classification"] = df_final[
                "Germline classification"
            ].replace(
                {
                    "Likely Pathogenic": "Likely pathogenic",
                    "Uncertain Significance": "Uncertain significance",
                    "Likely Benign": "Likely benign",
                }
            )
            if (df_final.Interpreted == "yes").sum() > 0 and list(
                df_final["Ref genome"].unique()
            )[0] != "not_defined":
                df_clinvar = df_final[df_final["Interpreted"] == "yes"]
                df_clinvar = df_clinvar[
                    [
                        "Local ID",
                        "Linking ID",
                        "Organisation ID",
                        "Gene symbol",
                        "Chromosome",
                        "Start",
                        "Reference allele",
                        "Alternate allele",
                        "Preferred condition name",
                        "Germline classification",
                        "Date last evaluated",
                        "Comment on classification",
                        "Collection method",
                        "Allele origin",
                        "Affected status",
                        "Ref genome",
                        "HGVSc",
                        "Consequence",
                        "Interpreted",
                        "Instrument ID",
                        "Specimen ID",
                    ]
                ]
                df_clinvar.to_csv(
                    arguments.outdir
                    + Path(filename).stem
                    + "_clinvar_variants.csv",
                    index=False,
                )
                write_txt_file(
                    arguments.clinvar_file_log,
                    filename,
                    "",
                )
                if not no_dx_upload:
                    dx_login(arguments.token)
                    now = datetime.now()
                    print("uploading clinvar csv to DNAnexus")
                    if clinvar_count == 0:
                        folder_name = (
                            "csvs_"
                            + now.strftime("%Y%m%d")
                            + "_"
                            + now.strftime("%H%M%S")
                        )
                        project = dxpy.DXProject(
                            config_variable["info"]["csv_projectID"]
                        )
                        project.new_folder(
                            folder=arguments.subfolder
                            + folder_name)
                    dxpy.upload_local_file(
                            arguments.outdir
                            + Path(filename).stem
                            + "_clinvar_variants.csv",
                            project=config_variable["info"]["csv_projectID"],
                            folder=arguments.subfolder
                            + folder_name
                        )
                clinvar_count = clinvar_count + 1
            elif list(df_final["Ref genome"].unique())[0] == "not_defined":
                write_txt_file(
                    arguments.failed_file_log,
                    filename,
                    "Ref_genome_not_defined",
                )
        df_final.to_csv(
            arguments.outdir + Path(filename).stem + "_all_variants.csv",
            index=False,
        )
        write_txt_file(arguments.parsed_file_log, filename, "")
        print("Successfully parsed", filename)
        shutil.move(filename, arguments.completed_dir)

    # uploading log files to dnanexus project for backup
    pf_base_name = Path(arguments.parsed_file_log).stem
    cf_base_name = Path(arguments.clinvar_file_log).stem
    now = datetime.now()
    if not no_dx_upload:
        print("uploading log file(s) to DNAnexus")
        dx_login(arguments.token)
        dxpy.upload_local_file(
            arguments.parsed_file_log,
            project=config_variable["info"]["csv_projectID"],
            folder="/parser_logs/",
            name=pf_base_name
            + "_"
            + now.strftime("%Y%m%d")
            + "_"
            + now.strftime("%H%M%S")
            + ".txt",
        )
    if not no_dx_upload and os.path.isfile(arguments.clinvar_file_log):
        dxpy.upload_local_file(
            arguments.clinvar_file_log,
            project=config_variable["info"]["csv_projectID"],
            folder="/parser_logs/",
            name=cf_base_name
            + "_"
            + now.strftime("%Y%m%d")
            + "_"
            + now.strftime("%H%M%S")
            + ".txt",
        )
    print("Done")


if __name__ == "__main__":
    main()
